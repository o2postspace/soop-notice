#!/usr/bin/env python3
"""구형 SOOPNOTICE 삼국지 Excel 워크북 생성기.

운영 원장은 Apps Script ``setupSamgukSheet()``가 관리한다. 이 파일의 A:AD
로컬 XLSX 스키마는 호환 보관용이며 운영 Google Sheet에 다시 올리면 안 된다.
"""

from __future__ import annotations

import argparse
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


MAX_INPUT_ROW = 5001
FONT_NAME = "Noto Sans CJK KR"
SHEET_URL = "https://docs.google.com/spreadsheets/d/1xC3leW9fFl4ytHI6i2UkQ8iViBFIwjLrug66lYmVckY/edit"
WIKI_URL = "https://threekingdoms.notion.site/"
SKILL_STATS_URL = "https://threekingdoms.notion.site/872f648a81d1821ea42581f61a4fc57e"
FMKOREA_RULES_URL = "https://www.fmkorea.com/10143176987"
FMKOREA_SLIDES_URL = "https://www.fmkorea.com/10143088032"
FACTIONS_URL = "https://gamcom-3kingdom.vercel.app/factions"
TERRITORY_URL = "https://gamcom-3kingdom.vercel.app/simulation"

COLORS = {
    "navy": "172033",
    "indigo": "5B6AED",
    "light_indigo": "EEF0FF",
    "white": "FFFFFF",
    "text": "222733",
    "muted": "667085",
    "border": "D6D9E0",
    "green": "D9EAD3",
    "yellow": "FFF2CC",
    "red": "F4CCCC",
    "gray": "E7E6E6",
    "blue": "D9EAF7",
    "wei": "DCE6F1",
    "shu": "E2F0D9",
    "wu": "FCE4D6",
}

THIN = Side(style="thin", color=COLORS["border"])
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor=COLORS["navy"])
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color=COLORS["white"])
BODY_FONT = Font(name=FONT_NAME, size=10, color=COLORS["text"])
LINK_FONT = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")
NATION_FILLS = {
    "위": PatternFill("solid", fgColor=COLORS["wei"]),
    "촉": PatternFill("solid", fgColor=COLORS["shu"]),
    "오": PatternFill("solid", fgColor=COLORS["wu"]),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--html",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "public" / "index.html",
    )
    parser.add_argument("--snapshot", type=Path)
    parser.add_argument("--territory", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path.home() / "삼국지_방송_추적시트.xlsx",
    )
    parser.add_argument(
        "--allow-legacy-schema",
        action="store_true",
        help="운영 원장과 호환되지 않는 구형 로컬 XLSX 생성을 명시적으로 허용합니다.",
    )
    return parser.parse_args()


def read_roster(html_path: Path) -> list[dict]:
    html = html_path.read_text(encoding="utf-8")
    match = re.search(r"const SAMGUK_MEMBERS = \[([\s\S]*?)\n\]\.map", html)
    if not match:
        raise RuntimeError("SAMGUK_MEMBERS를 찾지 못했습니다.")

    row_pattern = re.compile(
        r"\['([^']*)','([^']*)','([^']*)','([^']*)',(null|'[^']*')\]"
    )
    rows = []
    for index, row in enumerate(row_pattern.findall(match.group(1)), start=1):
        job = "" if row[4] == "null" else row[4][1:-1]
        rows.append(
            {
                "player_id": f"P{index:03d}",
                "nation": row[0],
                "crew": row[1],
                "nickname": row[2],
                "soop_id": row[3],
                "job": job,
            }
        )

    if len(rows) != 90:
        raise RuntimeError(f"참가자 수가 90명이 아닙니다: {len(rows)}")
    if len({row["soop_id"] for row in rows}) != len(rows):
        raise RuntimeError("SOOP ID가 중복되었습니다.")
    return rows


def load_snapshot(path: Path | None) -> tuple[dict[str, dict], dict]:
    if not path or not path.exists():
        return {}, {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, list):
        members, meta = raw, {}
    else:
        members = raw.get("members") or raw.get("data") or raw.get("players") or []
        meta = {key: value for key, value in raw.items() if key not in {"members", "data", "players"}}
    snapshot = {}
    for member in members:
        nickname = member.get("nickname") or member.get("name")
        if nickname:
            snapshot[str(nickname)] = member
    return snapshot, meta


def load_territories(path: Path | None) -> list[dict]:
    if not path or not path.exists():
        return []

    raw = json.loads(path.read_text(encoding="utf-8"))
    rows = []
    if isinstance(raw, list):
        rows = raw
    elif isinstance(raw, dict) and isinstance(raw.get("forces"), dict):
        for force_rows in raw["forces"].values():
            if isinstance(force_rows, list):
                rows.extend(force_rows)
    elif isinstance(raw, dict):
        for key in ("territories", "castles", "data", "rows"):
            if isinstance(raw.get(key), list):
                rows = raw[key]
                break

    territories = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        number_raw = item.get("number", item.get("name"))
        try:
            number = int(number_raw)
        except (TypeError, ValueError):
            continue
        owner = str(item.get("owner") or "미점령").strip()
        if owner not in {"위", "촉", "오", "미점령"}:
            owner = "미점령"
        territories.append(
            {
                "id": str(item.get("id") or item.get("castleKey") or f"영토-{number:02d}"),
                "number": number,
                "x": item.get("x"),
                "y": item.get("y"),
                "owner": owner,
                "capital": bool(item.get("capital", item.get("isCapital", False))),
                "facility": str(item.get("facility") or item.get("facilityType") or "없음"),
                "level": item.get("level"),
                "special": number == 27,
                "capture_status": "미점령" if owner == "미점령" else "점령",
                "capture_rate": item.get("captureRate", item.get("capture_rate")),
            }
        )

    territories.sort(key=lambda row: row["number"])
    if len(territories) != 60:
        raise RuntimeError(f"영토 수가 60개가 아닙니다: {len(territories)}")
    if {row["number"] for row in territories} != set(range(1, 61)):
        raise RuntimeError("영토 번호가 1~60과 일치하지 않습니다.")
    if len({row["id"] for row in territories}) != 60:
        raise RuntimeError("영토 ID가 중복되었습니다.")
    return territories


def style_header(ws, row: int, columns: int) -> None:
    for cell in ws[row][:columns]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER
    ws.row_dimensions[row].height = 31


def style_body(ws, min_row: int, max_row: int, columns: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=columns):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center")
            cell.border = GRID_BORDER


def set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def add_list_validation(ws, cell_range: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "목록에서 값을 선택하세요."
    validation.errorTitle = "잘못된 값"
    validation.prompt = "드롭다운 목록에서 선택할 수 있습니다."
    validation.promptTitle = "입력 안내"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_whole_validation(ws, cell_range: str, max_value: int) -> None:
    validation = DataValidation(
        type="whole", operator="between", formula1="0", formula2=str(max_value), allow_blank=True
    )
    validation.error = f"0~{max_value} 범위의 정수만 입력하세요."
    validation.errorTitle = "잘못된 숫자"
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_decimal_validation(ws, cell_range: str, max_value: int) -> None:
    validation = DataValidation(
        type="decimal", operator="between", formula1="0", formula2=str(max_value), allow_blank=True
    )
    validation.error = f"0~{max_value} 범위의 숫자만 입력하세요."
    validation.errorTitle = "잘못된 숫자"
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(cell_range)


def snapshot_value(member: dict, *keys):
    for key in keys:
        if key in member and member[key] is not None:
            return member[key]
    return None


def nonzero_snapshot_value(member: dict, *keys):
    value = snapshot_value(member, *keys)
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value if value else None
    if isinstance(value, (int, float)):
        return value if value != 0 else None
    if isinstance(value, str):
        try:
            return None if float(value.replace(",", "")) == 0 else value
        except ValueError:
            return value
    return value


def normalize_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "live", "on", "1", "yes"}:
            return True
        if lowered in {"false", "offline", "off", "0", "no"}:
            return False
    return None


def add_guide_sheet(
    wb: Workbook, generated_at: datetime, has_snapshot: bool, has_territory: bool
) -> None:
    ws = wb.create_sheet("사용법")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "삼국지 방송 추적 운영 시트"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(name=FONT_NAME, size=18, bold=True, color=COLORS["white"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 44

    instructions = [
        ("생성 시각", generated_at.strftime("%Y-%m-%d %H:%M:%S KST")),
        ("운영 원칙", "확인한 화면마다 완전한 스냅샷 한 행을 추가합니다. 기존 행을 덮어쓰거나 삭제하지 않습니다."),
        ("1. 방송 확인", "방송모니터링에서 LIVE 링크를 열고 레벨·말·장비·능력치를 확인합니다."),
        ("2. 값 기록", "관측입력에는 현재 스냅샷 전체와 근거 URL/타임코드·확인시각·증거해시를 함께 남깁니다."),
        ("3. 교차검증", "서로 다른 두 출처가 같은 값을 확인하거나 방송의 서로 다른 두 프레임이 일치한 스냅샷만 추가합니다."),
        (
            "4. 현황/파워랭킹",
            "현재현황은 교차검증된 최신 스냅샷 한 행을 통째로 반영합니다. 사이트 파워 v1은 레벨·기량 30%, 장비 강화 35%, 각인 20%, 말 강화 15%를 고정 가중치로 합산합니다.",
        ),
        (
            "5. 랭킹 신뢰도",
            "결측치는 0점이 아닌 0~100 가능 구간으로 남겨 전원을 중간값 추정순위에 표시합니다. 수집률 85%는 신뢰도 플래그이며, 100% 수집·현황/장비 교차검증·비교 표본 임계치를 모두 충족해야 확정입니다.",
        ),
        ("6. 영토 기록", "영토 변화도 교차검증된 완전한 스냅샷을 영토입력에 추가하고 가장 최근 확인시각의 행을 표시합니다."),
        (
            "OCR 연결",
            "OCR설정에 캡처 해상도와 항목별 좌표를 넣습니다. 최대체력은 플레이어 HUD의 현재/최대 표기에서 수집하고, 평타는 동일 대상·조건의 표본 수와 대표값을 함께 기록합니다. 두 전투 관측값은 파워 v1에 넣지 않습니다.",
        ),
        ("초기 데이터", "공개 지통실 조회값을 시트 기준값으로 넣었습니다." if has_snapshot else "현재 네트워크 스냅샷 없이 90명 명단만 넣었습니다."),
        ("초기 영토", "공개 지도 60개를 시트 기준값으로 넣었습니다." if has_territory else "영토 원본 없이 빈 입력 구조만 만들었습니다."),
        ("주의", "후보값은 이 시트에 넣지 않습니다. 교차검증 프로그램이 승격한 완전한 스냅샷만 아래에 추가하세요."),
    ]
    for row_index, (label, text) in enumerate(instructions, start=3):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, text)
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
        ws.cell(row_index, 1).font = Font(name=FONT_NAME, bold=True, color=COLORS["indigo"])
        ws.cell(row_index, 2).font = BODY_FONT
        ws.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="center")
        for cell in ws[row_index][:8]:
            cell.border = GRID_BORDER
        ws.row_dimensions[row_index].height = 34

    legend_row = 16
    ws.cell(legend_row, 1, "색상")
    ws.cell(legend_row, 2, "교차검증")
    ws.cell(legend_row, 3, "시트 기준값")
    ws.cell(legend_row, 4, "충돌/오류")
    ws.cell(legend_row, 5, "오래됨")
    for cell in ws[legend_row][:5]:
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.border = GRID_BORDER
        cell.alignment = Alignment(horizontal="center")
    ws.cell(legend_row, 2).fill = PatternFill("solid", fgColor=COLORS["green"])
    ws.cell(legend_row, 3).fill = PatternFill("solid", fgColor=COLORS["yellow"])
    ws.cell(legend_row, 4).fill = PatternFill("solid", fgColor=COLORS["red"])
    ws.cell(legend_row, 5).fill = PatternFill("solid", fgColor=COLORS["gray"])
    set_widths(ws, [18, 22, 22, 22, 22, 18, 18, 18])
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = COLORS["indigo"]


def add_reference_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("기준정보")
    groups = {
        1: ("국가", ["위", "촉", "오"]),
        3: ("활동상태", ["활동", "휴식", "하차"]),
        5: ("근거종류", [
            "시트", "에펨코리아", "방송", "시트+에펨코리아", "시트+방송",
            "에펨코리아+방송", "시트+에펨코리아+방송",
        ]),
        7: ("검증상태", ["기준값", "교차검증", "방송교차검증", "충돌"]),
        9: ("모니터링상태", ["대기", "수동확인", "OCR연결", "중지"]),
        11: ("방송상태", ["확인필요", "LIVE", "OFFLINE", "확인실패"]),
        16: ("영토소유", ["위", "촉", "오", "미점령"]),
        18: ("예/아니오", ["Y", "N"]),
        20: ("시설", ["없음", "병영", "성채", "장원"]),
        22: ("점령상태", ["미점령", "점령", "점령중", "교전", "확인필요"]),
    }
    for column, (header, values) in groups.items():
        ws.cell(1, column, header)
        for row, value in enumerate(values, start=2):
            ws.cell(row, column, value)
    ws["M1"] = "설정"
    ws["N1"] = "값"
    settings = [
        ("레벨 임시 상한", 10000),
        ("강화 임시 상한", 999),
        ("능력치 임시 상한", 1000000),
        ("OCR 신뢰도 상한", 100),
        ("현황 최신 기준(시간)", 1),
        ("재확인 기준(시간)", 6),
    ]
    for row, (label, value) in enumerate(settings, start=2):
        ws.cell(row, 13, label)
        ws.cell(row, 14, value)
    ws["M9"] = "운영 Google Sheet"
    ws["N9"] = SHEET_URL
    ws["M10"] = "SOOPNOTICE"
    ws["N10"] = "https://soopnotice.com"
    for cell in (ws["N9"], ws["N10"]):
        cell.hyperlink = cell.value
        cell.font = LINK_FONT
    style_header(ws, 1, 22)
    style_body(ws, 2, 10, 22)
    set_widths(ws, [17, 3, 17, 3, 19, 3, 17, 3, 19, 3, 17, 3, 24, 48, 3, 17, 3, 17, 3, 17, 3, 17])
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "8497B0"


def add_game_info_sheet(wb: Workbook, generated_at: datetime) -> None:
    ws = wb.create_sheet("게임정보")
    headers = ["분류", "항목", "내용", "출처URL", "기준일", "검수상태"]
    ws.append(headers)
    checked_on = generated_at.strftime("%Y-%m-%d")
    rows = [
        (
            "서버 개요",
            "10일간 진행되는 RPG 영토전",
            "2026년 8월 1일 21시부터 8월 10일 21시까지 진행됩니다. 요괴 사냥 RPG와 영토 전쟁이 결합되며, 종료 시 가장 많은 영토를 가진 나라가 우승합니다.",
            WIKI_URL,
        ),
        (
            "기량",
            "무력 · 기민 · 기력 · 지모",
            "각 기량은 1단계마다 기량점수 1을 사용합니다. 네 수치의 합은 기량 투자량으로만 사용하며, 직업별 효율과 장비·말·각인 효과가 달라 종합 전투력으로 보지 않습니다.",
            SKILL_STATS_URL,
        ),
        (
            "장비",
            "무기와 두갑 · 흉갑 · 각갑",
            "무기는 공격력·힘·공격속도에 영향을 줍니다. 방어구는 두갑·흉갑·각갑으로 나뉘며 두갑은 군주만 보유합니다.",
            FMKOREA_SLIDES_URL,
        ),
        (
            "강화",
            "강화 재료와 보조권",
            "무기는 강화석·석재·금화, 방어구는 강화석·목재·금화를 사용합니다. 단계 하락 방지권과 확률 2·3·5·10배 증가권이 있습니다. 공개 자료에 정확한 단계별 성공률·상한은 없어 임의로 추정하지 않습니다.",
            FMKOREA_RULES_URL,
        ),
        (
            "각인",
            "장비별 각인석 최대 3개",
            "각 장비에는 각인석을 최대 3개까지 적용할 수 있고, 동일 효과는 합연산됩니다.",
            WIKI_URL,
        ),
        (
            "영토",
            "60개 영토와 인접 구매",
            "총 60개 영토가 있으며 초기 구매는 상하좌우로 인접한 영토만 가능합니다. 구매비는 50만 금화입니다.",
            WIKI_URL,
        ),
        (
            "영토",
            "수도 · 시설 · 27번 특수 영토",
            "초기 수도는 위 8번, 촉 42번, 오 47번입니다. 시설은 병영·성채·장원이 있고 장원은 국가당 최대 10개입니다. 27번 특수 영토는 보유국 인원의 공격력을 5% 높입니다.",
            WIKI_URL,
        ),
        (
            "점령전",
            "점령률 변화",
            "자국 인원만 있으면 점령률이 상승하고 양국 인원이 함께 있으면 유지되며 상대국 인원만 있으면 하락합니다. 진행 중 수치는 방송 확인 또는 수기 입력이 필요합니다.",
            WIKI_URL,
        ),
    ]
    for category, title, description, source_url in rows:
        ws.append([category, title, description, source_url, checked_on, "참고"])

    style_header(ws, 1, len(headers))
    style_body(ws, 2, len(rows) + 1, len(headers))
    for row in range(2, len(rows) + 2):
        ws.cell(row, 3).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row, 4).hyperlink = ws.cell(row, 4).value
        ws.cell(row, 4).font = LINK_FONT
        ws.row_dimensions[row].height = 58
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"
    ws.freeze_panes = "A2"
    set_widths(ws, [15, 28, 92, 52, 14, 14])
    ws.sheet_properties.tabColor = "8064A2"


def add_participants_sheet(wb: Workbook, roster: list[dict], snapshot: dict[str, dict]) -> None:
    ws = wb.create_sheet("참가자")
    headers = [
        "player_id", "국가", "세력/길드", "닉네임", "SOOP_ID", "장수/직업",
        "활동상태", "프로필URL", "방송URL", "메모",
    ]
    ws.append(headers)
    for member in roster:
        snap = snapshot.get(member["nickname"], {})
        external_job = snapshot_value(snap, "job") or ""
        job = external_job or member["job"]
        note = ""
        if member["job"] and external_job and member["job"] != external_job:
            note = f"직업 불일치: 로컬 {member['job']} / 공개현황 {external_job} (확인 필요)"
        profile_url = f"https://www.sooplive.com/{member['soop_id']}"
        live_url = f"https://play.sooplive.com/{member['soop_id']}"
        ws.append([
            member["player_id"], member["nation"], member["crew"], member["nickname"],
            member["soop_id"], job, "활동", profile_url, live_url, note,
        ])
    style_header(ws, 1, len(headers))
    style_body(ws, 2, len(roster) + 1, len(headers))
    for row in range(2, len(roster) + 2):
        ws.cell(row, 2).fill = NATION_FILLS[ws.cell(row, 2).value]
        for column in (8, 9):
            ws.cell(row, column).hyperlink = ws.cell(row, column).value
            ws.cell(row, column).font = LINK_FONT
    add_list_validation(ws, f"B2:B{len(roster) + 1}", "'기준정보'!$A$2:$A$4")
    add_list_validation(ws, f"G2:G{len(roster) + 1}", "'기준정보'!$C$2:$C$4")
    ws.conditional_formatting.add(
        f"A2:A{len(roster) + 1}",
        FormulaRule(formula=["COUNTIF($A:$A,A2)>1"], fill=PatternFill("solid", fgColor=COLORS["red"])),
    )
    ws.conditional_formatting.add(
        f"E2:E{len(roster) + 1}",
        FormulaRule(formula=["COUNTIF($E:$E,E2)>1"], fill=PatternFill("solid", fgColor=COLORS["red"])),
    )
    ws.auto_filter.ref = f"A1:J{len(roster) + 1}"
    ws.freeze_panes = "A2"
    set_widths(ws, [12, 9, 18, 16, 20, 16, 13, 40, 40, 48])
    ws.sheet_properties.tabColor = "70AD47"


def add_monitoring_sheet(
    wb: Workbook, roster: list[dict], snapshot: dict[str, dict], snapshot_meta: dict, generated_at: datetime
) -> None:
    ws = wb.create_sheet("방송모니터링")
    headers = [
        "player_id", "국가", "세력/길드", "닉네임", "SOOP_ID", "방송링크",
        "방송상태", "방송제목", "시청자", "모니터링상태", "화면/창번호",
        "해상도", "OCR프로필", "상태확인시각", "담당자", "메모",
    ]
    ws.append(headers)
    fetched_at = snapshot_meta.get("fetched_at") or snapshot_meta.get("observed_at")
    if not fetched_at and snapshot:
        fetched_at = generated_at.replace(tzinfo=None)
    for member in roster:
        snap = snapshot.get(member["nickname"], {})
        live = normalize_bool(snapshot_value(snap, "live", "is_live", "isLive"))
        live_status = "LIVE" if live is True else "OFFLINE" if live is False else "확인필요"
        title = snapshot_value(snap, "title", "broad_title", "broadcast_title") or ""
        viewers = snapshot_value(snap, "viewers", "viewer", "current_viewers")
        live_url = f"https://play.sooplive.com/{member['soop_id']}"
        ws.append([
            member["player_id"], member["nation"], member["crew"], member["nickname"],
            member["soop_id"], live_url, live_status, title, viewers, "대기", "", "", "default",
            fetched_at or "", "", "",
        ])
    style_header(ws, 1, len(headers))
    style_body(ws, 2, len(roster) + 1, len(headers))
    for row in range(2, len(roster) + 2):
        ws.cell(row, 2).fill = NATION_FILLS[ws.cell(row, 2).value]
        ws.cell(row, 6).hyperlink = ws.cell(row, 6).value
        ws.cell(row, 6).font = LINK_FONT
        ws.cell(row, 14).number_format = "yyyy-mm-dd hh:mm:ss"
    add_list_validation(ws, f"G2:G{len(roster) + 1}", "'기준정보'!$K$2:$K$5")
    add_list_validation(ws, f"J2:J{len(roster) + 1}", "'기준정보'!$I$2:$I$5")
    ws.conditional_formatting.add(
        f"G2:G{len(roster) + 1}",
        CellIsRule(operator="equal", formula=['"LIVE"'], fill=PatternFill("solid", fgColor=COLORS["green"])),
    )
    ws.conditional_formatting.add(
        f"G2:G{len(roster) + 1}",
        CellIsRule(operator="equal", formula=['"확인실패"'], fill=PatternFill("solid", fgColor=COLORS["red"])),
    )
    ws.auto_filter.ref = f"A1:P{len(roster) + 1}"
    ws.freeze_panes = "A2"
    set_widths(ws, [12, 9, 18, 16, 20, 38, 13, 48, 11, 16, 14, 13, 15, 20, 14, 42])
    ws.sheet_properties.tabColor = "00B0F0"


def add_observations_sheet(
    wb: Workbook, roster: list[dict], snapshot: dict[str, dict], snapshot_meta: dict, generated_at: datetime
) -> int:
    ws = wb.create_sheet("관측입력")
    headers = [
        "observation_id", "player_id", "확인시각", "근거종류", "근거(URL/타임코드)",
        "레벨", "말", "말강화", "무기강화", "두갑강화", "흉갑강화", "각갑강화",
        "무력", "기민", "기력", "지모", "무력점수", "교차검증수", "검증상태",
        "증거해시", "수집배치", "기록자", "OCR신뢰도", "메모", "최대체력",
        "평타피해대표값", "평타표본수", "평타대상", "전투조건", "입력시각",
    ]
    ws.append(headers)
    comments = {
        1: "행을 고유하게 식별합니다. 자동 생성값을 가급적 수정하지 마세요.",
        3: "방송 화면에서 실제로 확인한 시각입니다.",
        5: "방송 URL, 다시보기 URL과 타임코드 등 재검증 가능한 근거입니다.",
        17: "기량과 별개로 외부 랭킹이 제공한 점수만 기록합니다. 임의 계산값을 넣지 않습니다.",
        18: "서로 다른 출처의 일치 개수입니다. 방송은 서로 다른 프레임 두 장을 한 출처의 교차검증으로 봅니다.",
        19: "후보값은 넣지 않습니다. 기준값·교차검증·방송교차검증 스냅샷만 현재현황에 반영됩니다.",
        23: "0~100. 방송교차검증은 95 이상인 서로 다른 두 프레임이 필요합니다.",
        25: "플레이어 HUD의 현재/최대 체력 중 최대값입니다. 우측 하단 군마 체력과 구분합니다.",
        26: "동일 대상·동일 조건의 비치명 기본공격 표본으로 만든 대표값입니다. 단일 화면의 숫자를 확정값으로 쓰지 않습니다.",
        27: "평타피해대표값 계산에 실제로 사용한 표본 수입니다.",
        28: "평타 표본을 수집한 동일 대상 또는 대상군입니다.",
        29: "무기·버프·치명타 여부 등 피해 비교에 필요한 조건입니다.",
        30: "값을 시트에 입력한 시각입니다.",
    }
    for column, text in comments.items():
        ws.cell(1, column).comment = Comment(text, "SOOPNOTICE")

    sources = snapshot_meta.get("sources") if isinstance(snapshot_meta.get("sources"), dict) else {}
    source_url = snapshot_meta.get("factions_source") or snapshot_meta.get("source_factions")
    source_url = source_url or sources.get("factions") or FACTIONS_URL
    fetched_at = snapshot_meta.get("fetched_at") or snapshot_meta.get("observed_at")
    if not fetched_at:
        fetched_at = generated_at.replace(tzinfo=None)

    initial_count = 0
    if snapshot:
        for index, member in enumerate(roster, start=1):
            snap = snapshot.get(member["nickname"], {})
            if not snap:
                continue
            values = [
                nonzero_snapshot_value(snap, "level"),
                snapshot_value(snap, "horse"),
                nonzero_snapshot_value(snap, "horse_level", "horseLevel"),
                nonzero_snapshot_value(snap, "weapon"),
                nonzero_snapshot_value(snap, "helmet"),
                nonzero_snapshot_value(snap, "armor"),
                nonzero_snapshot_value(snap, "shoes"),
                nonzero_snapshot_value(snap, "stat_strength", "strength"),
                nonzero_snapshot_value(snap, "stat_agility", "agility"),
                nonzero_snapshot_value(snap, "stat_vitality", "vitality"),
                nonzero_snapshot_value(snap, "stat_intelligence", "intelligence"),
                nonzero_snapshot_value(snap, "power_score", "powerScore", "power"),
            ]
            combat_values = [
                nonzero_snapshot_value(snap, "max_health", "maxHealth", "maxHP", "최대체력"),
                nonzero_snapshot_value(
                    snap, "basic_attack_damage", "basicAttackDamage", "평타피해대표값"
                ),
                nonzero_snapshot_value(
                    snap, "basic_attack_sample_count", "basicAttackSampleCount", "평타표본수"
                ),
                snapshot_value(snap, "basic_attack_target", "basicAttackTarget", "평타대상"),
                snapshot_value(snap, "combat_conditions", "combatConditions", "전투조건"),
            ]
            if not any(value not in (None, "") for value in values + combat_values):
                continue
            ws.append([
                f"INIT-{generated_at:%Y%m%d}-{index:03d}", member["player_id"], fetched_at,
                "시트", source_url, *values, 1, "기준값", "", f"INIT-{generated_at:%Y%m%d}",
                "초기 1회 수집", "", "공개 지통실의 시작 기준 스냅샷.", *combat_values,
                fetched_at,
            ])
            initial_count += 1

    style_header(ws, 1, len(headers))
    style_body(ws, 2, MAX_INPUT_ROW, len(headers))
    for row in range(2, MAX_INPUT_ROW + 1):
        ws.cell(row, 3).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row, 30).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="F2F4F7")
        ws.cell(row, 1).protection = Protection(locked=True)

    add_list_validation(ws, f"B2:B{MAX_INPUT_ROW}", "'참가자'!$A$2:$A$91")
    add_list_validation(ws, f"D2:D{MAX_INPUT_ROW}", "'기준정보'!$E$2:$E$8")
    add_list_validation(ws, f"S2:S{MAX_INPUT_ROW}", "'기준정보'!$G$2:$G$5")
    add_whole_validation(ws, f"F2:F{MAX_INPUT_ROW}", 10000)
    add_whole_validation(ws, f"H2:H{MAX_INPUT_ROW}", 80)
    add_whole_validation(ws, f"I2:L{MAX_INPUT_ROW}", 15)
    add_whole_validation(ws, f"M2:P{MAX_INPUT_ROW}", 1000000)
    add_decimal_validation(ws, f"Q2:Q{MAX_INPUT_ROW}", 1000000)
    add_whole_validation(ws, f"R2:R{MAX_INPUT_ROW}", 10)
    add_whole_validation(ws, f"W2:W{MAX_INPUT_ROW}", 100)
    add_whole_validation(ws, f"Y2:Y{MAX_INPUT_ROW}", 1000000)
    add_decimal_validation(ws, f"Z2:Z{MAX_INPUT_ROW}", 1000000)
    add_whole_validation(ws, f"AA2:AA{MAX_INPUT_ROW}", 10000)

    status_range = f"S2:S{MAX_INPUT_ROW}"
    status_colors = {
        "교차검증": COLORS["green"], "방송교차검증": COLORS["green"],
        "기준값": COLORS["blue"], "충돌": COLORS["red"],
    }
    for status, color in status_colors.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{status}"'], fill=PatternFill("solid", fgColor=color)),
        )
    ws.conditional_formatting.add(
        f"A2:A{MAX_INPUT_ROW}",
        FormulaRule(
            formula=['AND(A2<>"",COUNTIF($A:$A,A2)>1)'],
            fill=PatternFill("solid", fgColor=COLORS["red"]),
        ),
    )
    ws.auto_filter.ref = f"A1:AD{MAX_INPUT_ROW}"
    ws.freeze_panes = "F2"
    set_widths(ws, [
        22, 12, 20, 16, 48, 10, 16, 11, 12, 12, 12, 12, 11, 11, 11,
        11, 12, 13, 18, 44, 18, 16, 13, 52, 13, 16, 13, 24, 36, 20,
    ])
    ws.sheet_properties.tabColor = "FFC000"
    return initial_count


def latest_snapshot_row_formula(
    *, sheet: str, key_column: str, status_column: str, timestamp_column: str, row: int
) -> str:
    """Return one latest accepted snapshot row so values and provenance never mix."""
    key_range = f"{sheet}!${key_column}$2:${key_column}${MAX_INPUT_ROW}"
    status_range = f"{sheet}!${status_column}$2:${status_column}${MAX_INPUT_ROW}"
    timestamp_range = f"{sheet}!${timestamp_column}$2:${timestamp_column}${MAX_INPUT_ROW}"
    row_range = f"ROW({sheet}!${key_column}$2:${key_column}${MAX_INPUT_ROW})"
    accepted = (
        f'(({status_range}="기준값")+({status_range}="교차검증")+'
        f'({status_range}="방송교차검증"))>0'
    )
    maximum = f'MAX(FILTER({timestamp_range},{key_range}=$A{row},{accepted}))'
    filtered = (
        f'FILTER({row_range},{key_range}=$A{row},{accepted},'
        f'{timestamp_range}={maximum})'
    )
    return f'=IFERROR(INDEX({filtered},ROWS({filtered})),"")'


def snapshot_value_formula(
    *, row: int, selected_row_column: str, sheet: str, source_column: str
) -> str:
    selected = f"${selected_row_column}{row}"
    value = f"INDEX({sheet}!${source_column}:${source_column},{selected})"
    return f'=IF({selected}="","",IF({value}="","",{value}))'


def add_current_sheet(wb: Workbook, roster: list[dict]) -> None:
    ws = wb.create_sheet("현재현황")
    headers = [
        "player_id", "국가", "세력/길드", "닉네임", "SOOP_ID", "장수/직업", "레벨", "말",
        "말강화", "무기강화", "두갑강화", "흉갑강화", "각갑강화", "장비총강화",
        "무력", "기민", "기력", "지모", "무력점수", "최종확인", "최근근거", "출처종류",
        "교차검증수", "검증상태", "신선도", "기량합계", "랭킹점수", "공동순위",
        "정렬순번", "선택원본행", "최대체력", "평타피해대표값", "평타표본수",
        "평타대상", "전투조건",
    ]
    ws.append(headers)
    source_columns = {
        7: "F", 8: "G", 9: "H", 10: "I", 11: "J", 12: "K", 13: "L",
        15: "M", 16: "N", 17: "O", 18: "P", 19: "Q", 20: "C", 21: "E",
        22: "D", 23: "R", 24: "S", 31: "Y", 32: "Z", 33: "AA", 34: "AB", 35: "AC",
    }
    for row_index, member in enumerate(roster, start=2):
        participant_row = row_index
        ws.cell(row_index, 1, member["player_id"])
        for column in range(2, 7):
            participant_column = get_column_letter(column)
            ws.cell(row_index, column, f"=참가자!{participant_column}{participant_row}")
        ws.cell(
            row_index,
            30,
            latest_snapshot_row_formula(
                sheet="관측입력", key_column="B", status_column="S",
                timestamp_column="C", row=row_index,
            ),
        )
        for column, source in source_columns.items():
            ws.cell(
                row_index,
                column,
                snapshot_value_formula(
                    row=row_index, selected_row_column="AD", sheet="관측입력", source_column=source
                ),
            )
        ws.cell(row_index, 14, f'=IF(COUNT(J{row_index}:M{row_index})=0,"",SUM(J{row_index}:M{row_index}))')
        ws.cell(
            row_index,
            25,
            f'=IF(T{row_index}="","미확인",IF(NOW()-T{row_index}<=1/24,"최신",IF(NOW()-T{row_index}<=6/24,"확인 필요","오래됨")))',
        )
        ws.cell(row_index, 26, f'=IF(COUNT(O{row_index}:R{row_index})=0,"",SUM(O{row_index}:R{row_index}))')
        ws.cell(row_index, 27, f'=IF(COUNT($S$2:$S$91)>0,S{row_index},O{row_index})')
        ws.cell(row_index, 28, f'=IF(OR(AA{row_index}="",AA{row_index}<=0),"",RANK.EQ(AA{row_index},$AA$2:$AA$91,0))')
        ws.cell(row_index, 29, f'=IF(OR(AA{row_index}="",AA{row_index}<=0),"",RANK.EQ(AA{row_index},$AA$2:$AA$91,0)+COUNTIF($AA$2:AA{row_index},AA{row_index})-1)')
    style_header(ws, 1, len(headers))
    style_body(ws, 2, len(roster) + 1, len(headers))
    for row in range(2, len(roster) + 2):
        ws.cell(row, 2).fill = NATION_FILLS[roster[row - 2]["nation"]]
        ws.cell(row, 20).number_format = "yyyy-mm-dd hh:mm:ss"
    status_colors = {
        "교차검증": COLORS["green"], "방송교차검증": COLORS["green"],
        "기준값": COLORS["blue"], "충돌": COLORS["red"],
    }
    for status, color in status_colors.items():
        ws.conditional_formatting.add(
            "X2:X91",
            CellIsRule(operator="equal", formula=[f'"{status}"'], fill=PatternFill("solid", fgColor=color)),
        )
    freshness_colors = {
        "최신": COLORS["green"], "확인 필요": COLORS["yellow"],
        "오래됨": COLORS["gray"], "미확인": COLORS["gray"],
    }
    for status, color in freshness_colors.items():
        ws.conditional_formatting.add(
            "Y2:Y91",
            CellIsRule(operator="equal", formula=[f'"{status}"'], fill=PatternFill("solid", fgColor=color)),
        )
    ws.column_dimensions["AC"].hidden = True
    ws.column_dimensions["AD"].hidden = True
    ws.auto_filter.ref = "A1:AI91"
    ws.freeze_panes = "G2"
    set_widths(ws, [
        12, 9, 18, 16, 20, 16, 9, 16, 10, 11, 11, 11, 11, 13, 10, 10, 10,
        10, 12, 20, 46, 14, 13, 18, 14, 11, 12, 11, 10, 11, 13, 16, 13, 24, 36,
    ])
    ws.sheet_properties.tabColor = "4472C4"


def add_ranking_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("무력랭킹")
    headers = [
        "공동순위", "player_id", "국가", "세력/길드", "닉네임", "장수/직업",
        "레벨", "랭킹점수", "기준", "무력점수", "무력", "기민", "기력", "지모",
        "장비총강화", "최종확인", "출처종류", "교차검증수", "검증상태", "근거", "정렬순번",
    ]
    ws.append(headers)
    source_columns = {
        1: "AB", 2: "A", 3: "B", 4: "C", 5: "D", 6: "F", 7: "G", 8: "AA",
        10: "S", 11: "O", 12: "P", 13: "Q", 14: "R", 15: "N", 16: "T",
        17: "V", 18: "W", 19: "X", 20: "U",
    }
    for row in range(2, 92):
        ws.cell(row, 21, row - 1)
        for column, source in source_columns.items():
            ws.cell(
                row,
                column,
                f'=IFERROR(INDEX(현재현황!${source}$2:${source}$91,MATCH($U{row},현재현황!$AC$2:$AC$91,0)),"")',
            )
        ws.cell(row, 9, '=IF(COUNT(현재현황!$S$2:$S$91)>0,"무력점수","무력 스탯")')
    style_header(ws, 1, len(headers))
    style_body(ws, 2, 91, len(headers))
    for row in range(2, 92):
        ws.cell(row, 16).number_format = "yyyy-mm-dd hh:mm:ss"
    ws.column_dimensions["U"].hidden = True
    ws.auto_filter.ref = "A1:T91"
    ws.freeze_panes = "A2"
    set_widths(ws, [11, 12, 9, 18, 16, 16, 10, 12, 12, 12, 10, 10, 10, 10, 14, 20, 14, 13, 18, 48, 10])
    ws.sheet_properties.tabColor = "C00000"


def add_territory_input_sheet(
    wb: Workbook, territories: list[dict], generated_at: datetime
) -> int:
    ws = wb.create_sheet("영토입력")
    excel_timestamp = generated_at.replace(tzinfo=None)
    headers = [
        "territory_observation_id", "영토ID", "확인시각", "근거종류", "근거(URL/타임코드)",
        "번호", "X", "Y", "소유국", "수도", "시설", "레벨", "특수지", "점령상태",
        "점령률", "검증상태", "교차검증수", "증거해시", "메모", "입력시각",
    ]
    ws.append(headers)
    for index, territory in enumerate(territories, start=1):
        ws.append(
            [
                f"TINIT-{generated_at:%Y%m%d}-{index:03d}",
                territory["id"],
                excel_timestamp,
                "시트",
                TERRITORY_URL,
                territory["number"],
                territory["x"],
                territory["y"],
                territory["owner"],
                "Y" if territory["capital"] else "N",
                territory["facility"],
                territory["level"],
                "Y" if territory["special"] else "N",
                territory["capture_status"],
                territory["capture_rate"],
                "기준값",
                1,
                "",
                "공개 지도 초기 기준 스냅샷.",
                excel_timestamp,
            ]
        )

    initial_count = len(territories)
    style_header(ws, 1, len(headers))
    style_body(ws, 2, MAX_INPUT_ROW, len(headers))
    for row in range(2, MAX_INPUT_ROW + 1):
        ws.cell(row, 3).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row, 20).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="F2F4F7")
        ws.cell(row, 1).protection = Protection(locked=True)
    for row in range(2, initial_count + 2):
        ws.cell(row, 5).hyperlink = ws.cell(row, 5).value
        ws.cell(row, 5).font = LINK_FONT

    add_list_validation(ws, f"B2:B{MAX_INPUT_ROW}", "'영토현황'!$A$2:$A$61")
    add_list_validation(ws, f"D2:D{MAX_INPUT_ROW}", "'기준정보'!$E$2:$E$8")
    add_list_validation(ws, f"I2:I{MAX_INPUT_ROW}", "'기준정보'!$P$2:$P$5")
    add_list_validation(ws, f"J2:J{MAX_INPUT_ROW}", "'기준정보'!$R$2:$R$3")
    add_list_validation(ws, f"K2:K{MAX_INPUT_ROW}", "'기준정보'!$T$2:$T$5")
    add_list_validation(ws, f"M2:M{MAX_INPUT_ROW}", "'기준정보'!$R$2:$R$3")
    add_list_validation(ws, f"N2:N{MAX_INPUT_ROW}", "'기준정보'!$V$2:$V$6")
    add_list_validation(ws, f"P2:P{MAX_INPUT_ROW}", "'기준정보'!$G$2:$G$5")
    add_whole_validation(ws, f"F2:F{MAX_INPUT_ROW}", 60)
    add_whole_validation(ws, f"G2:H{MAX_INPUT_ROW}", 10000)
    add_whole_validation(ws, f"L2:L{MAX_INPUT_ROW}", 999)
    add_whole_validation(ws, f"O2:O{MAX_INPUT_ROW}", 100)
    add_whole_validation(ws, f"Q2:Q{MAX_INPUT_ROW}", 10)

    for status, color in {
        "교차검증": COLORS["green"], "방송교차검증": COLORS["green"],
        "기준값": COLORS["blue"], "충돌": COLORS["red"],
    }.items():
        ws.conditional_formatting.add(
            f"P2:P{MAX_INPUT_ROW}",
            CellIsRule(operator="equal", formula=[f'"{status}"'], fill=PatternFill("solid", fgColor=color)),
        )
    ws.conditional_formatting.add(
        f"A2:A{MAX_INPUT_ROW}",
        FormulaRule(
            formula=['AND(A2<>"",COUNTIF($A:$A,A2)>1)'],
            fill=PatternFill("solid", fgColor=COLORS["red"]),
        ),
    )
    ws.auto_filter.ref = f"A1:T{MAX_INPUT_ROW}"
    ws.freeze_panes = "F2"
    set_widths(ws, [25, 14, 20, 16, 48, 9, 9, 9, 11, 9, 12, 9, 10, 13, 10, 14, 14, 20, 48, 20])
    ws.sheet_properties.tabColor = "ED7D31"
    return initial_count


def add_territory_current_sheet(wb: Workbook, territories: list[dict]) -> None:
    ws = wb.create_sheet("영토현황")
    headers = [
        "영토ID", "번호", "X", "Y", "소유국", "수도", "시설", "레벨", "특수지",
        "점령상태", "점령률", "최종확인", "근거", "출처종류", "교차검증수",
        "검증상태", "메모", "선택원본행",
    ]
    ws.append(headers)
    source_columns = {
        2: "F", 3: "G", 4: "H", 5: "I", 6: "J", 7: "K", 8: "L",
        9: "M", 10: "N", 11: "O", 12: "C", 13: "E", 14: "D", 15: "Q", 17: "S",
    }
    for row_index, territory in enumerate(territories, start=2):
        ws.cell(row_index, 1, territory["id"])
        ws.cell(
            row_index,
            18,
            latest_snapshot_row_formula(
                sheet="영토입력", key_column="B", status_column="P",
                timestamp_column="C", row=row_index,
            ),
        )
        for column, source in source_columns.items():
            ws.cell(
                row_index,
                column,
                snapshot_value_formula(
                    row=row_index, selected_row_column="R", sheet="영토입력", source_column=source
                ),
            )
        ws.cell(
            row_index,
            16,
            snapshot_value_formula(
                row=row_index, selected_row_column="R", sheet="영토입력", source_column="P"
            ),
        )

    style_header(ws, 1, len(headers))
    if territories:
        style_body(ws, 2, len(territories) + 1, len(headers))
    for row in range(2, len(territories) + 2):
        ws.cell(row, 12).number_format = "yyyy-mm-dd hh:mm:ss"
    if territories:
        for owner, color in {
            "위": COLORS["wei"], "촉": COLORS["shu"], "오": COLORS["wu"], "미점령": COLORS["gray"],
        }.items():
            ws.conditional_formatting.add(
                f"E2:E{len(territories) + 1}",
                CellIsRule(operator="equal", formula=[f'"{owner}"'], fill=PatternFill("solid", fgColor=color)),
            )
        for status, color in {
            "교차검증": COLORS["green"], "방송교차검증": COLORS["green"],
            "기준값": COLORS["blue"], "충돌": COLORS["red"],
        }.items():
            ws.conditional_formatting.add(
                f"P2:P{len(territories) + 1}",
                CellIsRule(operator="equal", formula=[f'"{status}"'], fill=PatternFill("solid", fgColor=color)),
            )
    ws.column_dimensions["R"].hidden = True
    ws.auto_filter.ref = f"A1:Q{len(territories) + 1}"
    ws.freeze_panes = "E2"
    set_widths(ws, [14, 9, 9, 9, 11, 9, 13, 9, 10, 13, 10, 20, 48, 14, 13, 18, 48, 11])
    ws.sheet_properties.tabColor = "548235"


def add_ocr_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("OCR설정")
    headers = ["profile_id", "항목", "x", "y", "width", "height", "기준해상도", "배율", "활성", "샘플경로", "메모"]
    ws.append(headers)
    fields = [
        ("level", "레벨"), ("horse", "말"), ("horse_level", "말강화"),
        ("weapon", "무기강화"), ("helmet", "두갑강화"), ("armor", "흉갑강화"),
        ("shoes", "각갑강화"), ("strength", "무력"), ("agility", "기민"),
        ("vitality", "기력"), ("intelligence", "지모"), ("powerScore", "무력점수"),
        ("maxHealth", "최대체력"), ("basicAttackDamage", "평타피해대표값"),
    ]
    for key, label in fields:
        ws.append(["default", label, "", "", "", "", "1920x1080", 1, "N", "", f"OCR key: {key}"])
    style_header(ws, 1, len(headers))
    style_body(ws, 2, len(fields) + 1, len(headers))
    add_list_validation(ws, f"I2:I{len(fields) + 1}", '"Y,N"')
    add_whole_validation(ws, f"C2:F{len(fields) + 1}", 10000)
    ws.auto_filter.ref = f"A1:K{len(fields) + 1}"
    ws.freeze_panes = "A2"
    set_widths(ws, [16, 16, 10, 10, 10, 10, 16, 10, 10, 34, 28])
    ws.sheet_properties.tabColor = "7030A0"


def add_change_log_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("변경로그")
    headers = ["change_id", "승격시각", "확인시각", "player_id", "항목", "이전값", "새값", "근거", "출처", "교차검증수", "observation_id"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    style_body(ws, 2, 501, len(headers))
    for row in range(2, 502):
        ws.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row, 3).number_format = "yyyy-mm-dd hh:mm:ss"
    ws["A2"] = "후보는 로컬 append-only 큐에 보관하고, 교차검증으로 최신값이 승격된 이벤트만 이 시트에 기록합니다."
    ws.merge_cells("A2:K2")
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["yellow"])
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.auto_filter.ref = "A1:K501"
    ws.freeze_panes = "A2"
    set_widths(ws, [20, 20, 20, 12, 16, 14, 14, 48, 14, 14, 22])
    ws.sheet_properties.tabColor = "A5A5A5"


def build_workbook(
    html_path: Path,
    snapshot_path: Path | None,
    territory_path: Path | None,
    output_path: Path,
) -> dict:
    roster = read_roster(html_path)
    snapshot, snapshot_meta = load_snapshot(snapshot_path)
    territories = load_territories(territory_path)
    # 서버의 로컬 timezone(Asia/Seoul)을 사용해 Python 3.8에서도 동작시킨다.
    generated_at = datetime.now().astimezone().replace(microsecond=0)

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "SOOPNOTICE 삼국지 방송 추적 시트"
    wb.properties.creator = "SOOPNOTICE"
    wb.properties.description = "90명 방송 모니터링, 장비·기량·전투 관측 교차검증, 60개 영토 현황 및 파워랭킹 보조 원장"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    add_guide_sheet(wb, generated_at, bool(snapshot), bool(territories))
    add_game_info_sheet(wb, generated_at)
    add_reference_sheet(wb)
    add_participants_sheet(wb, roster, snapshot)
    add_monitoring_sheet(wb, roster, snapshot, snapshot_meta, generated_at)
    initial_count = add_observations_sheet(wb, roster, snapshot, snapshot_meta, generated_at)
    add_current_sheet(wb, roster)
    add_ranking_sheet(wb)
    territory_initial_count = add_territory_input_sheet(wb, territories, generated_at)
    add_territory_current_sheet(wb, territories)
    add_ocr_sheet(wb)
    add_change_log_sheet(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    check = load_workbook(output_path, read_only=False, data_only=False)
    expected = [
        "사용법", "게임정보", "기준정보", "참가자", "방송모니터링", "관측입력",
        "현재현황", "무력랭킹", "영토입력", "영토현황", "OCR설정", "변경로그",
    ]
    if check.sheetnames != expected:
        raise RuntimeError(f"시트 구성이 다릅니다: {check.sheetnames}")
    if check["참가자"].max_row != 91:
        raise RuntimeError("참가자 행 수 검증 실패")
    observation_headers = [check["관측입력"].cell(1, column).value for column in range(1, 31)]
    if observation_headers[24:30] != [
        "최대체력", "평타피해대표값", "평타표본수", "평타대상", "전투조건", "입력시각",
    ]:
        raise RuntimeError("관측입력 A:AD 전투 관측 스키마 검증 실패")
    current_headers = [check["현재현황"].cell(1, column).value for column in range(1, 36)]
    if current_headers[30:35] != [
        "최대체력", "평타피해대표값", "평타표본수", "평타대상", "전투조건",
    ]:
        raise RuntimeError("현재현황 AE:AI 전투 관측 스키마 검증 실패")
    if territories and check["영토현황"].max_row != 61:
        raise RuntimeError("영토 행 수 검증 실패")
    if territories:
        territory_ids = [check["영토현황"].cell(row, 1).value for row in range(2, 62)]
        if len(set(territory_ids)) != 60:
            raise RuntimeError("영토 ID 고유성 검증 실패")
        baseline_count = sum(
            check["영토입력"].cell(row, 16).value == "기준값" for row in range(2, 62)
        )
        if baseline_count != 60:
            raise RuntimeError(f"영토 초기 기준값 행 검증 실패: {baseline_count}")
        special_rows = [
            row for row in range(2, 62)
            if check["영토입력"].cell(row, 13).value == "Y"
        ]
        if len(special_rows) != 1 or check["영토입력"].cell(special_rows[0], 6).value != 27:
            raise RuntimeError("27번 특수 영토 검증 실패")
    seed_zero_cells = []
    for row in range(2, min(initial_count, 90) + 2):
        for column in range(8, 18):
            if check["관측입력"].cell(row, column).value == 0:
                seed_zero_cells.append(check["관측입력"].cell(row, column).coordinate)
    if seed_zero_cells:
        raise RuntimeError(f"초기 강화/기량 0값 정규화 실패: {seed_zero_cells[:5]}")
    for sheet_name, seeded_rows in (
        ("관측입력", initial_count),
        ("영토입력", territory_initial_count),
    ):
        first_blank_row = seeded_rows + 2
        ghost_ids = [
            check[sheet_name].cell(row, 1).coordinate
            for row in range(first_blank_row, MAX_INPUT_ROW + 1)
            if check[sheet_name].cell(row, 1).value not in (None, "")
        ]
        if ghost_ids:
            raise RuntimeError(f"{sheet_name} 빈 행 ID 사전 생성 검증 실패: {ghost_ids[:5]}")
    current_formula = str(check["현재현황"]["G2"].value or "")
    current_evidence_formula = str(check["현재현황"]["U2"].value or "")
    current_row_formula = str(check["현재현황"]["AD2"].value or "")
    current_combat_formulas = [str(check["현재현황"].cell(2, column).value or "") for column in range(31, 36)]
    growth_formula = str(check["현재현황"]["Z2"].value or "")
    growth_rank_formula = str(check["현재현황"]["AB2"].value or "")
    ranking_score_formula = str(check["무력랭킹"]["H2"].value or "")
    territory_formula = str(check["영토현황"]["E2"].value or "") if territories else ""
    territory_row_formula = str(check["영토현황"]["R2"].value or "") if territories else ""
    if not all(status in current_row_formula for status in ("기준값", "교차검증", "방송교차검증")):
        raise RuntimeError("현재현황 교차검증 스냅샷 선택 수식 검증 실패")
    if territories and not all(status in territory_row_formula for status in ("기준값", "교차검증", "방송교차검증")):
        raise RuntimeError("영토현황 교차검증 스냅샷 선택 수식 검증 실패")
    if "FILTER(" not in current_row_formula or "MAX(FILTER(" not in current_row_formula:
        raise RuntimeError("현재현황 최신 확인시각 수식 검증 실패")
    if "$AD2" not in current_formula or "$AD2" not in current_evidence_formula:
        raise RuntimeError("현재현황 값/근거 원본행 일치 검증 실패")
    for formula, source_column in zip(current_combat_formulas, ("Y", "Z", "AA", "AB", "AC")):
        if "$AD2" not in formula or f"관측입력!${source_column}:${source_column}" not in formula:
            raise RuntimeError("현재현황 전투 관측 원본행 일치 검증 실패")
    if territories and "$R2" not in territory_formula:
        raise RuntimeError("영토현황 원본행 일치 검증 실패")
    if "SUM(O2:R2)" not in growth_formula:
        raise RuntimeError("현재현황 기량합계 수식 검증 실패")
    if "RANK.EQ(AA2,$AA$2:$AA$91,0)" not in growth_rank_formula:
        raise RuntimeError("현재현황 무력 공동순위 수식 검증 실패")
    if "현재현황!$AA$2:$AA$91" not in ranking_score_formula:
        raise RuntimeError("무력랭킹 점수 연결 수식 검증 실패")
    check.close()

    with zipfile.ZipFile(output_path) as archive:
        bad_member = archive.testzip()
        if bad_member is not None:
            raise RuntimeError(f"XLSX ZIP 손상: {bad_member}")
        required_members = {"[Content_Types].xml", "xl/workbook.xml"}
        if not required_members.issubset(set(archive.namelist())):
            raise RuntimeError("XLSX ZIP 필수 파일 검증 실패")

    return {
        "output": str(output_path),
        "members": len(roster),
        "snapshot_members": len(snapshot),
        "initial_observations": initial_count,
        "territories": len(territories),
        "territory_initial_observations": territory_initial_count,
        "sheets": expected,
        "formula_qa": "latest_cross_verified_snapshot",
        "zip_qa": "ok",
        "generated_at": generated_at.isoformat(),
    }


def main() -> None:
    args = parse_args()
    if not args.allow_legacy_schema:
        raise SystemExit(
            "구형 A:AD XLSX 생성기는 운영 원장과 호환되지 않아 차단되었습니다. "
            "운영 원장은 Apps Script setupSamgukSheet()를 사용하세요. "
            "보관용 XLSX가 꼭 필요하면 --allow-legacy-schema를 명시하세요."
        )
    result = build_workbook(args.html, args.snapshot, args.territory, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
